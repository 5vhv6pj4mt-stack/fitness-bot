"""Импортирует все тренировки из Excel в БД."""
import asyncio
import re
import openpyxl
from datetime import datetime
import aiosqlite
from config import DB_PATH

EXCEL_PATH = "/home/oracle/fitness-bot/ТРЕНИРОВКИ С ИИ (1).xlsx"
KIRILL_ID = 311739548

SKIP_SHEETS = {"ТЕХНИКА", "Лист19", "PROGRESS", "05.01.26old"}

WEEK_TYPE_MAP = {
    "СИЛОВАЯ": "strength", "СИЛОВОЙ": "strength",
    "ОБЪЁМНАЯ": "volume", "ОБЪЕМНАЯ": "volume", "ОБЪЁМ": "volume", "ОБЪЕМ": "volume",
    "РАЗГРУЗОЧНАЯ": "deload", "РАЗГРУЗОЧНЫЙ": "deload",
}

DAY_TYPE_MAP = {
    "ВЕРХ-СИЛА": "upper_strength", "ВЕРХ СИЛА": "upper_strength", "ВЕРХ-СИЛ": "upper_strength",
    "НОГИ": "legs",
    "ВЕРХ-ОБЪЕМ": "upper_volume", "ВЕРХ-ОБЪЁМ": "upper_volume",
    "ВЕРХ ОБЪЕМ": "upper_volume", "ВЕРХ ОБЪЁМ": "upper_volume",
}


def normalize(text: str) -> str:
    return re.sub(r'\s+', ' ', str(text).upper().strip())


def parse_val(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", ".").strip()
    try:
        return float(s)
    except ValueError:
        return None


def detect_week_type(text: str) -> str:
    t = normalize(text)
    for key, val in WEEK_TYPE_MAP.items():
        if key in t:
            return val
    return None


def detect_day_type(text: str) -> str:
    t = normalize(text)
    for key, val in DAY_TYPE_MAP.items():
        if key in t:
            return val
    return None


def detect_format(header_row) -> str:
    """Определяет формат листа по строке заголовка.
    format1: Вес 1 | Повт 1 | RPE 1 | Вес 2 ... (Dec 2025)
    format2a: Подходы | Вес(П) | Факт 1 | RPE 1 | Факт 2 ... (05.01.26, только повторения)
    format2b: Подходы | Вес(П) | Вес(Ф) | Повт(Ф) | RPE 1 | Вес(Ф) ... (Jan-May 2026)
    """
    cols = [str(v).lower() if v else "" for v in header_row[:6]]
    if any("вес 1" in c for c in cols):
        return "format1"
    if any("факт" in c for c in cols[3:]):
        return "format2a"
    return "format2b"


def parse_sets(row, fmt: str) -> list[dict]:
    """Парсит подходы из строки упражнения в зависимости от формата."""
    sets_data = []
    if fmt == "format1":
        # col1=Вес1, col2=Повт1, col3=RPE1, col4=Вес2, ...
        for s in range(4):
            base = 1 + s * 3
            if base + 2 >= len(row):
                break
            w = parse_val(row[base])
            r = parse_val(row[base + 1])
            rpe = parse_val(row[base + 2])
            if w is not None and r is not None and r > 0:
                sets_data.append({
                    "set_number": s + 1,
                    "planned_weight": w,
                    "actual_weight": w,
                    "reps": int(r),
                    "rpe": rpe if rpe and 1 <= rpe <= 10 else 7.0,
                })
    elif fmt == "format2a":
        # col1=ПланПодх, col2=ПланВес, col3=Факт1(повт), col4=RPE1, col5=Факт2, ...
        planned_w = parse_val(row[2]) if len(row) > 2 else None
        if planned_w is None or planned_w == 0:
            return []
        for s in range(4):
            base = 3 + s * 2
            if base + 1 >= len(row):
                break
            r = parse_val(row[base])
            rpe = parse_val(row[base + 1])
            if r is not None and r > 0:
                sets_data.append({
                    "set_number": s + 1,
                    "planned_weight": planned_w,
                    "actual_weight": planned_w,
                    "reps": int(r),
                    "rpe": rpe if rpe and 1 <= rpe <= 10 else 7.0,
                })
    else:
        # format2b: col1=ПланПодх, col2=ПланВес, col3=Вес(Ф)1, col4=Повт(Ф)1, col5=RPE1, col6=Вес(Ф)2, ...
        planned_w = parse_val(row[2]) if len(row) > 2 else None
        for s in range(4):
            base = 3 + s * 3
            if base + 2 >= len(row):
                break
            w = parse_val(row[base])
            r = parse_val(row[base + 1])
            rpe = parse_val(row[base + 2])
            if w is not None and r is not None and r > 0:
                sets_data.append({
                    "set_number": s + 1,
                    "planned_weight": planned_w or w,
                    "actual_weight": w,
                    "reps": int(r),
                    "rpe": rpe if rpe and 1 <= rpe <= 10 else 7.0,
                })
    return sets_data


def parse_sheet(ws) -> list[dict]:
    """Возвращает список сессий тренировок из листа."""
    sessions = []
    current_session = None
    current_week_type = None
    current_week_num = None
    header_found = False
    fmt = "format2b"

    def save_session():
        if current_session and current_session.get("exercises"):
            sessions.append(current_session)

    for row in ws.iter_rows(values_only=True):
        if all(v is None for v in row):
            continue

        first = str(row[0]).strip() if row[0] is not None else ""

        # Разделитель между сессиями (Format 3)
        if first == "-" and all(str(v) in ("-", "None", "") for v in row[:5]):
            save_session()
            current_session = None
            header_found = False
            continue

        # Итоги дня — сохраняем сессию и сбрасываем
        if "ИТОГИ" in first.upper() or "ОБЩИЕ" in first.upper():
            save_session()
            current_session = None
            header_found = False
            continue

        # Заголовок с номером недели и типом
        if "НЕДЕЛЯ" in first.upper():
            wt = detect_week_type(first)
            if wt:
                current_week_type = wt
            m = re.search(r'НЕДЕЛЯ\s+(\d+)', first.upper())
            current_week_num = int(m.group(1)) if m else current_week_num
            continue

        # Дата
        if first.startswith("Дата:"):
            date_val = row[1]
            if isinstance(date_val, datetime):
                if current_session is None:
                    current_session = {
                        "week_type": current_week_type,
                        "week_num": current_week_num,
                        "exercises": [],
                    }
                current_session["date"] = date_val.strftime("%Y-%m-%d")
            continue

        # День тренировки
        if first.startswith("День"):
            day_text = str(row[1]) if row[1] else first
            dt = detect_day_type(day_text) or detect_day_type(first)
            if current_session is None:
                current_session = {
                    "week_type": current_week_type,
                    "week_num": current_week_num,
                    "exercises": [],
                }
            if dt:
                current_session["day_type"] = dt
            continue

        # Строка заголовка таблицы
        if first.startswith("Упражнение") or first == "УПРАЖНЕНИЕ":
            fmt = detect_format(row)
            header_found = True
            continue

        # Строка с упражнением
        if header_found and current_session is not None and first and not first.startswith("-"):
            exercise_name = first.split("(")[0].strip()
            exercise_name = re.sub(r'\d+м\d*с.*$', '', exercise_name).strip()
            sets_data = parse_sets(row, fmt)
            if sets_data:
                current_session["exercises"].append({
                    "name": exercise_name,
                    "sets": sets_data,
                })

    save_session()
    return sessions


async def import_all():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    all_sessions = []
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        sessions = parse_sheet(ws)
        for s in sessions:
            if not s.get("date"):
                # Попробуем взять из имени листа
                m = re.match(r'(\d{2})\.(\d{2})\.(\d{2})', sheet_name)
                if m:
                    d, mo, y = m.groups()
                    s["date"] = f"20{y}-{mo}-{d}"
            if s.get("date") and s.get("exercises"):
                all_sessions.append(s)

    print(f"Найдено сессий: {len(all_sessions)}")

    async with aiosqlite.connect(DB_PATH) as db:
        inserted = 0
        skipped = 0

        for s in sorted(all_sessions, key=lambda x: x.get("date", "")):
            date = s["date"]
            day_type = s.get("day_type", "unknown")
            week_type = s.get("week_type") or "strength"
            week_num = s.get("week_num") or 0

            # Проверяем дубликат
            async with db.execute(
                "SELECT id FROM workouts WHERE user_id=? AND date=? AND day_type=?",
                (KIRILL_ID, date, day_type)
            ) as cur:
                if await cur.fetchone():
                    skipped += 1
                    continue

            # Считаем тоннаж и avg RPE
            all_sets = [st for ex in s["exercises"] for st in ex["sets"]]
            tonnage = sum(st["actual_weight"] * st["reps"] for st in all_sets)
            avg_rpe = sum(st["rpe"] for st in all_sets) / len(all_sets) if all_sets else 0

            cur = await db.execute(
                "INSERT INTO workouts (user_id, date, day_type, week_number, week_type, total_tonnage, avg_rpe) VALUES (?,?,?,?,?,?,?)",
                (KIRILL_ID, date, day_type, week_num, week_type, tonnage, avg_rpe)
            )
            workout_id = cur.lastrowid

            for ex in s["exercises"]:
                for st in ex["sets"]:
                    await db.execute(
                        "INSERT INTO workout_sets (workout_id, exercise, set_number, planned_weight, actual_weight, reps, rpe) VALUES (?,?,?,?,?,?,?)",
                        (workout_id, ex["name"], st["set_number"], st["planned_weight"],
                         st["actual_weight"], st["reps"], st["rpe"])
                    )
            inserted += 1

        await db.commit()
        print(f"Загружено: {inserted} тренировок, пропущено дублей: {skipped}")

        # Показываем что загрузили
        async with db.execute(
            "SELECT date, day_type, week_type, total_tonnage FROM workouts WHERE user_id=? ORDER BY date",
            (KIRILL_ID,)
        ) as cur:
            rows = await cur.fetchall()
            print(f"\nВсего тренировок в БД: {len(rows)}")
            for r in rows[-5:]:
                print(f"  {r[0]} | {r[2]} | {r[1]} | тоннаж: {r[3]:.0f}кг")


if __name__ == "__main__":
    asyncio.run(import_all())
